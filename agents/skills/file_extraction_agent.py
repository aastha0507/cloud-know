"""Agent for extracting content from various file types."""
from typing import Dict, Any, Optional
import mimetypes
from pathlib import Path
import io


class FileExtractionAgent:
    """Agent responsible for extracting text content from files."""
    
    def __init__(self):
        """Initialize the file extraction agent."""
        self.supported_types = {
            "text/plain": self._extract_text,
            "text/csv": self._extract_csv,
            "text/markdown": self._extract_text,
            "application/json": self._extract_json,
            "application/pdf": self._extract_pdf,
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document": self._extract_docx,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": self._extract_xlsx,
        }
    
    def extract(self, file_content: bytes, mime_type: str, file_name: Optional[str] = None) -> Dict[str, Any]:
        """Extract text content from a file.
        
        Args:
            file_content: File content as bytes
            mime_type: MIME type of the file
            file_name: Optional file name for context
            
        Returns:
            Dictionary with extracted content and metadata
        """
        try:
            # Try to get extractor for this MIME type
            extractor = self.supported_types.get(mime_type)
            
            if extractor:
                content = extractor(file_content)
            else:
                # Try to extract as text for unknown types
                try:
                    content = file_content.decode("utf-8")
                except UnicodeDecodeError:
                    content = f"[Binary file: {mime_type}]"
            
            return {
                "content": content,
                "mime_type": mime_type,
                "file_name": file_name,
                "extraction_method": "success"
            }
        except Exception as e:
            return {
                "content": f"[Error extracting content: {str(e)}]",
                "mime_type": mime_type,
                "file_name": file_name,
                "extraction_method": "error",
                "error": str(e)
            }
    
    def _extract_text(self, content: bytes) -> str:
        """Extract text from plain text files."""
        return content.decode("utf-8")
    
    def _extract_csv(self, content: bytes) -> str:
        """Extract text from CSV files."""
        import csv
        text_content = content.decode("utf-8")
        reader = csv.reader(io.StringIO(text_content))
        lines = []
        for row in reader:
            lines.append(" | ".join(row))
        return "\n".join(lines)
    
    def _extract_json(self, content: bytes) -> str:
        """Extract text from JSON files."""
        import json
        data = json.loads(content.decode("utf-8"))
        return json.dumps(data, indent=2)
    
    def _extract_pdf(self, content: bytes) -> str:
        """Extract text from PDF files."""
        try:
            import PyPDF2
            pdf_reader = PyPDF2.PdfReader(io.BytesIO(content))
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
            return text
        except ImportError:
            return "[PDF extraction requires PyPDF2 library]"
        except Exception as e:
            return f"[Error extracting PDF: {str(e)}]"
    
    def _extract_docx(self, content: bytes) -> str:
        """Extract text from DOCX files."""
        try:
            from docx import Document
            doc = Document(io.BytesIO(content))
            text = []
            for paragraph in doc.paragraphs:
                text.append(paragraph.text)
            return "\n".join(text)
        except ImportError:
            return "[DOCX extraction requires python-docx library]"
        except Exception as e:
            return f"[Error extracting DOCX: {str(e)}]"
    
    def _extract_xlsx(self, content: bytes) -> str:
        """Extract text from XLSX files."""
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(io.BytesIO(content))
            text = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text.append(f"Sheet: {sheet_name}")
                for row in sheet.iter_rows(values_only=True):
                    text.append(" | ".join(str(cell) if cell else "" for cell in row))
            return "\n".join(text)
        except ImportError:
            return "[XLSX extraction requires openpyxl library]"
        except Exception as e:
            return f"[Error extracting XLSX: {str(e)}]"

